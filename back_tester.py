import numpy as np
import pandas as pd
import csv
import matplotlib.pyplot as plt
from itertools import compress
import datetime as dt
from datetime import date
import urllib.request
import football_model as fm
import xlsxwriter
from openpyxl import Workbook
from openpyxl import load_workbook

#------------------------------------------------------------------------------------#

# Back Tester:

def team_query_outcome_at_date(dataframe,team,date):
    h_team_df = dataframe[dataframe['HomeTeam'] == team]
    h_team_df.reset_index(drop=True,inplace=True)
    h_team_df = h_team_df[h_team_df['Date'] < date]
    h_team = h_team_df.to_numpy()
    h_team = h_team[-fm.games_played_var:,3:5]
   
    a_team_df = dataframe[dataframe['AwayTeam'] == team]
    a_team_df.reset_index(drop=True,inplace=True)
    a_team_df = a_team_df[a_team_df['Date'] < date]
    a_team = a_team_df.to_numpy()
    a_team = a_team[-fm.games_played_var:,3:5]
    a_team = a_team[:,[1,0]]



    win_outcome = fm.win_candidate(fm.win_points_ratio(h_team),fm.goal_ratio(h_team))
    lose_outcome = fm.lose_candidate(fm.win_points_ratio(a_team),fm.goal_ratio(a_team))


    return win_outcome, lose_outcome

def league_candidates_at_date(league_key,date,yr):

    df = pd.read_csv('https://www.football-data.co.uk/mmz4281/'+ yr +'/' + league_key + '.csv')
    df = df[['Date','HomeTeam','AwayTeam','FTHG','FTAG']]
    df['Date'] = pd.to_datetime(df['Date'],dayfirst=True)
    teams_list = fm.get_team_list('https://www.football-data.co.uk/mmz4281/'+ yr +'/'+ league_key + '.csv')
    results = np.zeros([int(len(teams_list)),2])
    for i in range(int(len(teams_list))):
        results[i,:] = team_query_outcome_at_date(df,teams_list[i],date)
    

    return [list(compress(teams_list,results[:,0])),list(compress(teams_list,results[:,1]))]

def get_weeks_fixtures_at_date(league_key,date,yr):
    df = pd.read_csv('https://www.football-data.co.uk/mmz4281/'+ yr +'/'+ league_key + '.csv')
    df = df[['Date','HomeTeam','AwayTeam','B365H']]
    df['Date'] = pd.to_datetime(df['Date'],dayfirst=True).dt.date
    plus_week = dt.timedelta(days = 7)
    #df.to_csv('output1.csv')
    df = df[(date <= df['Date']) & (df['Date'] < date+plus_week)]
    df = df[['Date','HomeTeam','AwayTeam','B365H','B365H']]

    
    fixtures = df.to_numpy()

    
    return fixtures

def get_bets(league_key,date,yr):
    fixture_list = []
    candidates = league_candidates_at_date(league_key,date,yr)
    for i in range(len(get_weeks_fixtures_at_date(league_key,date,yr)[:,0])):
            if get_weeks_fixtures_at_date(league_key,date,yr)[i,1] in candidates[0]:
                if get_weeks_fixtures_at_date(league_key,date,yr)[i,2] in candidates[1]:
                    fixture_list.append([get_weeks_fixtures_at_date(league_key,date,yr)[i,0],get_weeks_fixtures_at_date(league_key,date,yr)[i,1],get_weeks_fixtures_at_date(league_key,date,yr)[i,2],
                        get_weeks_fixtures_at_date(league_key,date,yr)[i,3],get_weeks_fixtures_at_date(league_key,date,yr)[i,4]])
              
    return  fixture_list

def change_odds(league_key,date,yr):
    # changege odds on get_bets() fixtures output to 0 if game is loss leave if won.
    bets_list = get_bets(league_key,date,yr)
    df = pd.read_csv('https://www.football-data.co.uk/mmz4281/'+ yr +'/' + league_key + '.csv')
    df = df[['Date','HomeTeam','AwayTeam','FTHG','FTAG']]
    df['Date'] = pd.to_datetime(df['Date'],dayfirst=True)
    plus_week = dt.timedelta(days = 7)
    df = df[(date <= df['Date']) & (df['Date'] < date+plus_week)]
    df = df[['HomeTeam','AwayTeam','FTHG','FTAG']]
    fixtures = df.to_numpy()
    


    for i in range(len(fixtures[:,0])):
        for j in range(len(bets_list)):
            if (fixtures[i,0] == bets_list[j][1]) and (fixtures[i,1] == bets_list[j][2]):
                bets_list[j].append(str(fixtures[i,2])+'-'+str(fixtures[i,3]))
                if fixtures[i,2] > fixtures[i,3]:
                    bets_list[j][3] = bets_list[j][3]
                else:
                    bets_list[j][3] = 0

    return bets_list

def profit_loss_calc(stake,odds):
    profit = (stake*odds)-stake
        
    return round(profit,2)

def stake(bankroll):
    if bankroll <= 150:
        stake = 10
    elif 150 < bankroll <= 200:
        stake = 20
    else:
        stake =  30 

    return stake

def results_txt(listt):
    with open('back_test_results.txt', 'a+') as f:
        for item in listt:
            for j in item:
                f.write('%s     ' % j)
            f.write('\n')
    
    return

def write_out_excel(listt,yr):
   
    wb = load_workbook('results'+yr+'.xlsx')
    sheet = wb.active
    for row in listt:
        sheet.append(row)

    wb.save(filename='results'+yr+'.xlsx')

    return

def initialise_file(yr):
    listt = [['Date','Home Team','Away Team','Odds','Score','Result','Profit/Loss','Running Profit/Loss']]
    
    with xlsxwriter.Workbook('results'+yr+'.xlsx') as workbook:
        worksheet = workbook.add_worksheet('Results')

        for row_num, data in enumerate(listt):
            worksheet.write_row(row_num, 0, data)

    return

def select_date(yr):

    if yr =='1819':
        start_date = pd.to_datetime('05/12/2018',dayfirst=True)
        end_date = pd.to_datetime('06/05/2019',dayfirst=True)

        



    elif yr == '1718':
        start_date = pd.to_datetime('02/12/2017',dayfirst=True) 
        end_date = pd.to_datetime('07/05/2018',dayfirst=True)
        
        



    elif yr == '1617':
        #01/12/2016
        start_date = pd.to_datetime('24/02/2017',dayfirst=True)
        end_date = pd.to_datetime('14/05/2017',dayfirst=True)


        


    return start_date, end_date

def backtester_run(yr):

    bankroll = 100
    running_profit_loss = 0
    leagues_dic = {'EPL': 'E0', 'Championship': 'E1','Bundesliga 1': 'D1', 'Serie A': 'I1','La Liga Primera':'SP1', 'Le Championnat': 'F1'}
    leagues_dic2 = {'EPL':'E0'}
    initialise_file(yr)
    start_date,end_date = select_date(yr) 
    date = start_date

    while  date < end_date:
        for league_key in leagues_dic:
            bets_list = change_odds(leagues_dic[league_key],date,yr)
            for i in range(len(bets_list)): 
                bets_list[i].append(profit_loss_calc(stake(bankroll),bets_list[i][3]))
                bankroll = bankroll + profit_loss_calc(stake(bankroll),bets_list[i][3])
                running_profit_loss = running_profit_loss + profit_loss_calc(stake(bankroll),bets_list[i][3])
                bets_list[i].append(round(running_profit_loss,2))
                if bets_list[i][3] != 0:
                    bets_list[i][3] = 'WIN'
                else:
                    bets_list[i][3] = 'LOSS'

                bets_list[i][:] = [bets_list[i][j] for j in [0,1,2,4,5,3,6,7]]
            write_out_excel(bets_list,yr)
        
        date = date + dt.timedelta(days = 7)

    return




backtester_run('1617')








#------------------------------------------------------------------------------------#

# checking teams are correclty identified in win class/loss class.

def get_last_5_home(league_key,date,yr,team):
    date = pd.to_datetime(date,dayfirst=True)
    df = pd.read_csv('https://www.football-data.co.uk/mmz4281/'+ yr +'/'+ league_key + '.csv')
    df = df[['Date','HomeTeam','AwayTeam','FTHG','FTAG']]
    df['Date'] = pd.to_datetime(df['Date'],dayfirst=True).dt.date
    df = df[df['Date'] <= date]
    df = df[df['HomeTeam'] == team]
    fixtures = df.to_numpy()
    print(fixtures[-5:,:])
    fixtures = fixtures[-5:,3:5]
    print(fm.win_points_ratio(fixtures))
    print(fm.goal_ratio(fixtures))
    return

def get_last_5_away(league_key,date,yr,team):
    date = pd.to_datetime(date,dayfirst=True)
    df = pd.read_csv('https://www.football-data.co.uk/mmz4281/'+ yr +'/'+ league_key + '.csv')
    df = df[['Date','HomeTeam','AwayTeam','FTHG','FTAG']]
    df['Date'] = pd.to_datetime(df['Date'],dayfirst=True).dt.date
    df = df[df['Date'] <= date]
    df = df[df['AwayTeam'] == team]
    fixtures = df.to_numpy()
    print(fixtures[-5:,:])
    fixtures = fixtures[-5:,3:5]
    fixtures = fixtures[:,[1,0]]
    print(fm.win_points_ratio(fixtures))
    print(fm.goal_ratio(fixtures))
    return


#get_last_5_home('E0','30/01/2017','1617','Burnley')




















